from PySide6.QtWidgets import QFileDialog, QWidget
from .ui_main_widget import Ui_mainWindow
import openpyxl
import os

class MainWidget(QWidget):
    ui_form: Ui_mainWindow

    def __init__(self, ui_form: Ui_mainWindow):
        super().__init__()
        self.ui_form = ui_form
        self.ui_form.setupUi(self)
        self.setupUI()
    
    def setupUI(self): 
        self.ui_form.fileExplorerButton.clicked.connect(lambda: self.fileExplorerButtonClick())
        self.ui_form.mergeFilesOperationButton.clicked.connect(lambda: self.mergeFileOperationButtonClick())

    def fileExplorerButtonClick(self):
        dir = QFileDialog.getExistingDirectory(self, 'Выберите папку', options=QFileDialog.Option.ShowDirsOnly)
        self.ui_form.fileExplorerInput.setText(dir)

    def mergeFileOperationButtonClick(self): 
        filenames = [ os.path.join(self.ui_form.fileExplorerInput.toPlainText(), f) 
            for f in os.listdir(self.ui_form.fileExplorerInput.toPlainText()) 
            if os.path.isfile(os.path.join(self.ui_form.fileExplorerInput.toPlainText(), f))]
        wb = openpyxl.Workbook()
        ws = wb.create_sheet('Лист 1')
        header_flag = True
        for filename in filenames:
            temp_wb = openpyxl.load_workbook(filename)
            temp_ws = temp_wb.active
            if temp_ws: 
                if header_flag:
                    for i in range(1, temp_ws.max_column):
                        ws.cell(row=1, column=i, value=temp_ws.cell(row=1, column=i))
                    header_flag = False

            
